import os
import json
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelService:
    def __init__(self):
        self.excel_file = 'jobs.xlsx'
        self.json_file = 'jobs.json'
    
    def save_job(self, job_data):
        """Save a single job to both JSON and Excel"""
        try:
            # Save to JSON first
            self.save_to_json(job_data)
            
            # Save to Excel
            self.save_to_excel(job_data)
            
            logging.info("Job successfully saved to Excel and JSON")
            return True
        
        except Exception as e:
            logging.error(f"Error saving job to Excel: {str(e)}")
            return False
    
    def save_to_json(self, job_data):
        """Save job to JSON file"""
        try:
            # Load existing jobs
            if os.path.exists(self.json_file):
                with open(self.json_file, 'r', encoding='utf-8') as f:
                    jobs = json.load(f)
            else:
                jobs = []
            
            # Add new job
            job_data['id'] = len(jobs) + 1
            if 'posted_date' not in job_data:
                job_data['posted_date'] = datetime.utcnow().isoformat()
            
            jobs.append(job_data)
            
            # Save back to file
            with open(self.json_file, 'w', encoding='utf-8') as f:
                json.dump(jobs, f, indent=2, ensure_ascii=False)
            
            logging.info("Job saved to JSON file")
        
        except Exception as e:
            logging.error(f"Error saving to JSON: {str(e)}")
            raise
    
    def save_to_excel(self, job_data):
        """Save job to Excel file"""
        try:
            if os.path.exists(self.excel_file):
                workbook = load_workbook(self.excel_file)
                worksheet = workbook.active
            else:
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Jobs"
                
                # Create headers
                headers = [
                    'ID', 'Title', 'Company', 'Location', 'Category', 'Job Type',
                    'Experience', 'Salary', 'Description', 'Requirements',
                    'Application URL', 'Contact Email', 'Posted Date', 'Deadline'
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # Add new job row
            next_row = worksheet.max_row + 1
            
            job_row = [
                job_data.get('id', next_row - 1),
                job_data.get('title', ''),
                job_data.get('company', ''),
                job_data.get('location', ''),
                job_data.get('category', ''),
                job_data.get('job_type', ''),
                job_data.get('experience', ''),
                job_data.get('salary', ''),
                job_data.get('description', ''),
                job_data.get('requirements', ''),
                job_data.get('application_url', ''),
                job_data.get('contact_email', ''),
                job_data.get('posted_date', datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')),
                job_data.get('deadline', '')
            ]
            
            for col, value in enumerate(job_row, 1):
                worksheet.cell(row=next_row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(self.excel_file)
            logging.info("Job saved to Excel file")
        
        except Exception as e:
            logging.error(f"Error saving to Excel: {str(e)}")
            raise
    
    def get_all_jobs(self):
        """Get all jobs from JSON file"""
        try:
            if os.path.exists(self.json_file):
                with open(self.json_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return []
        
        except Exception as e:
            logging.error(f"Error reading jobs from JSON: {str(e)}")
            return []
    
    def export_all_jobs(self, jobs_data):
        """Export all jobs to a new Excel file"""
        try:
            export_file = f'jobs_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "All Jobs"
            
            # Create headers
            headers = [
                'ID', 'Title', 'Company', 'Location', 'Category', 'Job Type',
                'Experience', 'Salary', 'Description', 'Requirements',
                'Application URL', 'Contact Email', 'Posted Date', 'Deadline', 'Status'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add job data
            for row_idx, job in enumerate(jobs_data, 2):
                job_row = [
                    job.get('id', ''),
                    job.get('title', ''),
                    job.get('company', ''),
                    job.get('location', ''),
                    job.get('category', ''),
                    job.get('job_type', ''),
                    job.get('experience', ''),
                    job.get('salary', ''),
                    job.get('description', ''),
                    job.get('requirements', ''),
                    job.get('application_url', ''),
                    job.get('contact_email', ''),
                    job.get('posted_date', ''),
                    job.get('deadline', ''),
                    'Active' if job.get('is_active', True) else 'Inactive'
                ]
                
                for col, value in enumerate(job_row, 1):
                    worksheet.cell(row=row_idx, column=col, value=value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(export_file)
            logging.info(f"All jobs exported to {export_file}")
            return export_file
        
        except Exception as e:
            logging.error(f"Error exporting to Excel: {str(e)}")
            raise
